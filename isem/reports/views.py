# views.py
from django.http import HttpResponse
from django.utils.dateparse import parse_date
from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, Image, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4

from datetime import timedelta

import json

from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Count, Sum, F, Q
from django.db.models.functions import TruncDay, TruncMonth
from django.shortcuts import render
from django.utils.timezone import now

import base64
from io import BytesIO
from reportlab.lib.utils import ImageReader
from django.contrib.staticfiles import finders

from appointment.models import Appointment
from patient.models import Patient
from billing.models import BillingRecord
from inventory.models import InventoryItem


@login_required
@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def reports_dashboard(request):
    # -----------------------------
    # Config & Filters
    # -----------------------------
    range_type = request.GET.get("range", "daily")  # daily | monthly
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    today = now().date()
    start_date = today - timedelta(days=30)
    end_date = today

    # Determine Date Range
    if start_date_str and start_date_str != "None" and start_date_str.strip():
        parsed_start = parse_date(start_date_str)
        if parsed_start:
            start_date = parsed_start

    if end_date_str and end_date_str != "None" and end_date_str.strip():
        parsed_end = parse_date(end_date_str)
        if parsed_end:
            end_date = parsed_end

    trunc = TruncMonth if range_type == "monthly" else TruncDay
    date_format = "%b %Y" if range_type == "monthly" else "%b %d"

    # -----------------------------
    # Summary Metrics
    # -----------------------------
    completed_appointments = Appointment.objects.filter(
        status="done",
        date__range=[start_date, end_date]
    ).count()

    total_patients = Patient.objects.count()

    paid_revenue = (
        BillingRecord.objects.filter(
            payment_status="paid",
            date_issued__range=[start_date, end_date]
        ).aggregate(total=Sum("amount"))["total"] or 0
    )

    low_stock_count = InventoryItem.objects.filter(
        stock__lte=F("low_stock_threshold")
    ).count()

    expired_item_count = InventoryItem.objects.filter(
        status="expired"
    ).count()

    # -----------------------------
    # Appointment Trend
    # -----------------------------
    appointment_trend = (
        Appointment.objects.filter(
            date__range=[start_date, end_date],
            status__in=["arrived", "ongoing", "done"]
        )
        .annotate(period=trunc("date"))
        .values("period")
        .annotate(count=Count("id"))
        .order_by("period")
    )

    appointment_labels = [a["period"].strftime(date_format) for a in appointment_trend]
    appointment_data = [a["count"] for a in appointment_trend]

    # -----------------------------
    # Patient Growth
    # -----------------------------
    patient_growth = (
        Patient.objects.filter(created_at__range=[start_date, end_date])
        .annotate(period=trunc("created_at"))
        .values("period")
        .annotate(count=Count("id"))
        .order_by("period")
    )

    patient_labels = [p["period"].strftime(date_format) for p in patient_growth]
    patient_data = [p["count"] for p in patient_growth]

    # -----------------------------
    # Revenue Trend (PAID ONLY)
    # -----------------------------
    revenue_trend = (
        BillingRecord.objects.filter(
            payment_status="paid",
            date_issued__range=[start_date, end_date]
        )
        .annotate(period=trunc("date_issued"))
        .values("period")
        .annotate(total=Sum("amount"))
        .order_by("period")
    )

    revenue_labels = [r["period"].strftime(date_format) for r in revenue_trend]
    revenue_data = [float(r["total"] or 0) for r in revenue_trend]

    # -----------------------------
    # Inventory by Category
    # -----------------------------
    inventory_by_category = (
        InventoryItem.objects
        .values("category")
        .annotate(count=Count("id"))
        .order_by("category")
    )

    inventory_category_labels = [i["category"] or "Uncategorized" for i in inventory_by_category]
    inventory_category_data = [i["count"] for i in inventory_by_category]

    # NEW: Report Table Logic
    report_type = request.GET.get('report_type', 'appointment')

    report_data = []
    headers = []

    # 1. Selection & Filtering Logic strictly following models.txt
    if report_type == 'appointment':
        qs = Appointment.objects.filter(date__range=[start_date, end_date])
        headers = ['ID', 'Date', 'Time', 'Dentist', 'Branch', 'Status'] # [cite: 6, 8, 10]
        for obj in qs:
            report_data.append([
                obj.display_id, obj.date, obj.time, 
                obj.dentist.name if obj.dentist else obj.dentist_name, # [cite: 9, 11]
                obj.branch.name if obj.branch else "N/A", obj.get_status_display()
            ])

    elif report_type == 'patient':
        qs = Patient.objects.filter(created_at__date__range=[start_date, end_date])
        headers = ['Name', 'Email', 'Age', 'Gender', 'Telephone', 'Guest'] # [cite: 22, 23]
        for obj in qs:
            report_data.append([
                obj.name, obj.email, obj.age, obj.gender, obj.telephone, "Yes" if obj.is_guest else "No"
            ])

    elif report_type == 'billing':
        qs = BillingRecord.objects.filter(date_issued__date__range=[start_date, end_date])
        headers = ['Patient', 'Type', 'Amount', 'Status', 'Date'] # [cite: 14]
        for obj in qs:
            report_data.append([
                obj.patient_name, obj.type, f"₱{obj.amount}", obj.get_payment_status_display(), obj.date_issued.date()
            ])

    elif report_type == 'inventory':
        qs = InventoryItem.objects.all()
        headers = ['Item', 'Category', 'Stock', 'Threshold', 'Status'] # [cite: 16, 17]
        for obj in qs:
            report_data.append([
                obj.item_name, obj.get_category_display(), obj.stock, 
                obj.low_stock_threshold, obj.get_status_display()
            ])

    # -----------------------------
    # Context
    # -----------------------------
    context = {
        "range_type": range_type,

        # Summary cards
        "completed_appointments": completed_appointments,
        "total_patients": total_patients,
        "paid_revenue": paid_revenue,
        "low_stock_count": low_stock_count,
        "expired_item_count": expired_item_count,

        # Charts
        "appointment_labels": json.dumps(appointment_labels),
        "appointment_data": json.dumps(appointment_data),

        "patient_labels": json.dumps(patient_labels),
        "patient_data": json.dumps(patient_data),

        "revenue_labels": json.dumps(revenue_labels),
        "revenue_data": json.dumps(revenue_data),

        "inventory_category_labels": json.dumps(inventory_category_labels),
        "inventory_category_data": json.dumps(inventory_category_data),

        # Table Reports
        "report_type": report_type,
        "headers": headers,
        "report_data": report_data,
        "start_date": start_date.strftime('%Y-%m-%d'),
        "end_date": end_date.strftime('%Y-%m-%d'),
    }

    return render(request, "reports/index.html", context)

# -----------------------------
# Shared Data Builder
# -----------------------------
def build_report_data(range_type, start_date, end_date):
    trunc = TruncMonth if range_type == "monthly" else TruncDay
    date_format = "%b %Y" if range_type == "monthly" else "%b %d"

    appointment_trend = (
        Appointment.objects
        .filter(date__range=[start_date, end_date], status__in=["arrived", "ongoing", "done"])
        .annotate(period=trunc("date"))
        .values("period")
        .annotate(count=Count("id"))
        .order_by("period")
    )

    patient_growth = (
        Patient.objects
        .filter(created_at__date__range=[start_date, end_date])
        .annotate(period=trunc("created_at"))
        .values("period")
        .annotate(count=Count("id"))
        .order_by("period")
    )

    revenue_trend = (
        BillingRecord.objects
        .filter(payment_status="paid", date_issued__date__range=[start_date, end_date])
        .annotate(period=trunc("date_issued"))
        .values("period")
        .annotate(total=Sum("amount"))
        .order_by("period")
    )

    inventory_summary = (
        InventoryItem.objects
        .values("category")
        .annotate(
            total=Count("id"),
            low_stock=Count("id", filter=F("status") == "low_stock"),
            expired=Count("id", filter=F("status") == "expired"),
        )
    )

    return {
        "date": now().date(),
        "start_date": start_date,
        "end_date": end_date,
        "range": range_type,
        "appointment_chart": [(a["period"].strftime(date_format), a["count"]) for a in appointment_trend],
        "patient_chart": [(p["period"].strftime(date_format), p["count"]) for p in patient_growth],
        "revenue_chart": [(r["period"].strftime(date_format), float(r["total"] or 0)) for r in revenue_trend],
        "inventory_chart": list(inventory_summary),
        "raw": {
            "appointments": Appointment.objects.filter(date__range=[start_date, end_date]),
            "patients": Patient.objects.filter(created_at__date__range=[start_date, end_date]),
            "billing": BillingRecord.objects.filter(date_issued__date__range=[start_date, end_date]),
            "inventory": InventoryItem.objects.all(),
        }
    }

# -----------------------------
# Export View (Excel / PDF)
# -----------------------------
@login_required
@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def export_reports_dashboard(request):
    range_type = request.GET.get("range", "daily")
    export_format = request.GET.get("format", "excel")
    
    start_str = request.GET.get('start_date')
    end_str = request.GET.get('end_date')
    
    today = now().date()

    # Parse inputs (returning None if "None" or empty)
    p_start = parse_date(start_str) if start_str and start_str != "None" else None
    p_end = parse_date(end_str) if end_str and end_str != "None" else None

    # --- THE SMART LOGIC ---
    if p_start and not p_end:
        # Only start date picked? Use it for both.
        start_date = p_start
        end_date = p_start 
    elif p_end and not p_start:
        # Only end date picked? Use it for both.
        start_date = p_end
        end_date = p_end
    elif p_start and p_end:
        # Both picked? Use both.
        start_date = p_start
        end_date = p_end
    else:
        # Neither picked? Default to 30-day window.
        start_date = today - timedelta(days=30)
        end_date = today

    # Pass these specific dates to the builder
    data = build_report_data(range_type, start_date, end_date)

    if export_format == "pdf":
        return export_pdf_report(data, request)
    return export_excel_report(data)

# -----------------------------
# PDF Renderer
# -----------------------------
def export_pdf_report(data, request):
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = (
        f'attachment; filename="Clinic_Report_{data["range"]}_{data["date"]}.pdf"'
    )

    doc = SimpleDocTemplate(
        response, 
        pagesize=A4, 
        rightMargin=40, leftMargin=40, topMargin=30, bottomMargin=30
    )
    styles = getSampleStyleSheet()
    elements = []

    # --- 1. CENTERED TOP BANNER (Logo over Title) ---
    logo_path = finders.find('landingpage/logo.png') # Ensure file is at static/images/logo.png
    
    if logo_path:
        # 1a. Logo
        logo = Image(logo_path, width=100, height=30)
        logo.hAlign = 'CENTER'
        elements.append(logo)
        elements.append(Spacer(1, 10))

    # 1b. Title (Centered)
    title_style = styles['Title']
    title_style.alignment = 1  # 1 is Center
    elements.append(Paragraph("CLINIC REPORTS SUMMARY", title_style))
    
    # 1c. Sub-header (Centered)
    sub_style = styles['Normal']
    sub_style.alignment = 1
    # Format the dates
    fmt_start = data['start_date'].strftime('%b %d, %Y')
    fmt_end = data['end_date'].strftime('%b %d, %Y')

    # Smart Label: If the dates are the same, just show "Date: [Date]"
    if fmt_start == fmt_end:
        period_text = f"Date: {fmt_start}"
    else:
        period_text = f"Period: {fmt_start} to {fmt_end}"

    elements.append(Paragraph(
        f"<font color='gray'>Performance Report ({data['range'].upper()})<br/>"
        f"{period_text}<br/>"
        f"Generated on: {data['date']}</font>", 
        sub_style
    ))
    
    # Blue Divider Line
    line_table = Table([['']], colWidths=[480])
    line_table.setStyle(TableStyle([
        ('LINEBELOW', (0,0), (-1,-1), 2, colors.green),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
    ]))
    elements.append(line_table)
    elements.append(Spacer(1, 25))

    # ---------------------------------------------------
    # CHART VISUALIZATION SECTION
    # ---------------------------------------------------
    def get_chart_cell(chart_id, label):
        """Helper to create a title and image for a table cell"""
        chart_base64 = request.POST.get(chart_id)
        cell_elements = []
        cell_elements.append(Paragraph(f"<b>{label}</b>", styles["Normal"]))
        
        if chart_base64 and "," in chart_base64:
            try:
                img_str = chart_base64.split(',')[1]
                img_data = BytesIO(base64.b64decode(img_str))
                # Square-ish dimensions for the grid
                img = Image(img_data, width=230, height=150) 
                cell_elements.append(img)
            except Exception:
                cell_elements.append(Paragraph("[Chart Error]", styles["Normal"]))
        else:
            cell_elements.append(Paragraph("[No Data]", styles["Normal"]))
        
        return cell_elements

    # Prepare the 2x2 grid data
    # Row 1: Appointment (Top Left) | Patient (Top Right)
    # Row 2: Revenue (Bottom Left)     | Inventory (Bottom Right)
    row1 = [get_chart_cell('appointmentChart', 'Appointments'), get_chart_cell('patientChart', 'New Patients')]
    row2 = [get_chart_cell('revenueChart', 'Revenue Trend'), get_chart_cell('inventoryChart', 'Inventory Levels')]

    chart_grid = Table([row1, row2], colWidths=[250, 250])
    chart_grid.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('LEFTPADDING', (0,0), (-1,-1), 5),
        ('RIGHTPADDING', (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 15),
    ]))
    
    elements.append(chart_grid)
    elements.append(Spacer(1, 10))

    # --- 3. HELPER FOR SUMMARY TABLES ---
    def create_summary_table(header_title, table_headers, table_data, col_widths):
        elements.append(Paragraph(f"<b>{header_title}</b>", styles["Heading3"]))
        elements.append(Spacer(1, 5))
        t = Table([table_headers] + table_data, colWidths=col_widths)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.blue),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTSIZE', (0,0), (-1,-1), 8),
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ('TOPPADDING', (0,0), (-1,-1), 6),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 15))

    # --- 4. THE DATA TABLES ---

    # Appointment Summary
    apt_data = [[str(a.display_id), str(a.date), a.status, a.dentist.name if a.dentist else "N/A"] for a in data["raw"]["appointments"][:10]]
    create_summary_table("Recent Appointments (Top 10)", ["ID", "Date", "Status", "Dentist"], apt_data, [80, 100, 100, 150])

    # Patient Growth Summary
    pat_data = [[p.name, p.email, str(p.created_at.date())] for p in data["raw"]["patients"][:10]]
    create_summary_table("New Patients Summary", ["Name", "Email", "Registration Date"], pat_data, [150, 180, 100])

    # Billing/Revenue Summary
    bill_data = [[b.patient.name, f"P {b.amount}", b.payment_status, str(b.date_issued.date())] for b in data["raw"]["billing"][:10]]
    create_summary_table("Financial Records (Recent)", ["Patient", "Amount", "Status", "Date"], bill_data, [150, 80, 100, 100])

    doc.build(elements)
    return response

# -----------------------------
# Excel Renderer
# -----------------------------
def export_excel_report(data):
    wb = Workbook()

    # Appointments
    ws = wb.active
    ws.title = "Appointments"
    ws.append(["ID", "Date", "Status", "Dentist", "Branch"])
    for a in data["raw"]["appointments"]:
        ws.append([
            a.display_id,
            a.date,
            a.status,
            a.dentist.name if a.dentist else "",
            a.branch.name if a.branch else ""
        ])

    # Patients
    ws = wb.create_sheet("Patients")
    ws.append(["Name", "Email", "Created At"])
    for p in data["raw"]["patients"]:
        ws.append([p.name, p.email, p.created_at.date()])

    # Billing
    ws = wb.create_sheet("Billing")
    ws.append(["Patient", "Amount", "Issued Date"])
    for b in data["raw"]["billing"]:
        ws.append([b.patient.name, float(b.amount), b.date_issued.date()])

    # Inventory
    ws = wb.create_sheet("Inventory")
    ws.append(["Item", "Category", "Stock", "Status"])
    for i in data["raw"]["inventory"]:
        ws.append([i.item_name, i.category, i.stock, i.status])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="Clinic_Data_{data["range"]}_{data["date"]}.xlsx"'
    )

    wb.save(response)
    return response